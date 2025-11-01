"""Excel 解析器"""

import io
import logging

import openpyxl

from app.core.parsers.base import BaseParser

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Excel 文档解析器"""

    async def parse(self, file_data: bytes, **_kwargs) -> str:
        """解析 Excel 文档"""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True)

            text_parts = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # 添加 sheet 标题
                text_parts.append(f"Sheet: {sheet_name}\n")

                # 提取行数据
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        text_parts.append(row_text)

                text_parts.append("\n")

            workbook.close()
            return "\n".join(text_parts)

        except Exception as e:
            logger.error(f"Failed to parse Excel: {e}")
            raise
